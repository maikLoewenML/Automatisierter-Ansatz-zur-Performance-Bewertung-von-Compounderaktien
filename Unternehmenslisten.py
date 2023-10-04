import os
import openpyxl


def lese_sp500_unternehmen(jahr):
    dateipfad = f"Unternehmenslisten_S&P500/S&P500_Unternehmen_{jahr}.xlsx"
    # Prüfen, ob die Datei existiert
    if not os.path.exists(dateipfad):
        print(f"Datei {dateipfad} nicht gefunden!")
        return []
    # Excel-Datei öffnen
    workbook = openpyxl.load_workbook(dateipfad)
    worksheet = workbook.active
    aktiensymbole = []
    # Durch alle Zeilen im Arbeitsblatt gehen, beginnend mit der zweiten Zeile
    for row in range(2, worksheet.max_row + 1):
        symbol = worksheet.cell(row=row, column=1).value
        if symbol:
            aktiensymbole.append(symbol)
    return aktiensymbole
